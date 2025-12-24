"""Helpers for parsing synthetic commitments data used in tests."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List

from docx import Document
from openpyxl import load_workbook


@dataclass(frozen=True)
class CommitmentRecord:
    """Represents a single commitments workbook entry."""

    identifier: str
    vendor: str
    phase: str
    amount: float


def _normalise_header(header: Iterable[object]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for index, value in enumerate(header):
        key = str(value or "").strip().lower()
        if key:
            mapping[key] = index
    return mapping


def read_commitments_workbook(path: Path) -> List[CommitmentRecord]:
    """Load the synthetic commitments workbook into dataclasses."""

    workbook = load_workbook(path)
    try:
        if not workbook.sheetnames:
            return []
        sheet = workbook[workbook.sheetnames[0]]
        
        row_iterator = sheet.iter_rows(values_only=True)
        
        # Read and process the header row
        try:
            header_row = next(row_iterator)
        except StopIteration:
            return []
        
        header = _normalise_header(header_row)
        required = {
            "commitment id": "identifier",
            "vendor": "vendor",
            "phase": "phase",
            "amount": "amount",
        }
        missing = [key for key in required if key not in header]
        if missing:
            raise KeyError(f"Workbook is missing required columns: {', '.join(sorted(missing))}")

        # Process data rows iteratively
        records: List[CommitmentRecord] = []
        for raw_row in row_iterator:
            if not any(raw_row):
                continue
            identifier = str(raw_row[header["commitment id"]]).strip()
            vendor = str(raw_row[header["vendor"]]).strip()
            phase = str(raw_row[header["phase"]]).strip()
            amount_raw = raw_row[header["amount"]]
            try:
                amount = float(amount_raw)
            except (TypeError, ValueError):
                amount = 0.0
            records.append(CommitmentRecord(identifier, vendor, phase, amount))
        return records
    finally:
        workbook.close()


def read_environment_document(path: Path) -> Dict[str, str]:
    """Parse key-value metadata out of the generated environment document."""

    document = Document(path)
    details: Dict[str, str] = {}
    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if not text or ":" not in text:
            continue
        key, value = text.split(":", 1)
        details[key.strip()] = value.strip()
    return details


def summarise_commitments(records: Iterable[CommitmentRecord]) -> Dict[str, object]:
    """Compute aggregate totals grouped by vendor."""

    totals_by_vendor: Dict[str, float] = {}
    total_amount = 0.0
    for record in records:
        totals_by_vendor.setdefault(record.vendor, 0.0)
        totals_by_vendor[record.vendor] += record.amount
        total_amount += record.amount
    return {
        "total_amount": total_amount,
        "by_vendor": totals_by_vendor,
    }


__all__ = [
    "CommitmentRecord",
    "read_commitments_workbook",
    "read_environment_document",
    "summarise_commitments",
]
